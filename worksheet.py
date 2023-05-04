import pandas as pd
import json
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from loader import ExcelSheetAutomator
import subprocess
from send2trash import send2trash


class SheetLoader(ExcelSheetAutomator):

    def __init__(self):
        super().__init__()
        self.create_sheet()
        self.data_loader()

    def create_sheet(self):
        """Function used for creating a new Excel sheet/ workbook"""
        wb = Workbook()
        wb.save(filename=self.file_name)

    def data_loader(self):
        """ Function used for rendering the dataframe to an actual Excel sheet."""
        wb = openpyxl.load_workbook(self.file_name)
        sheet = wb.active
        rows = dataframe_to_rows(self.json_reader())  # ExcelSheetAutomator.json_reader(self)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        sheet.delete_rows(idx=2)
        wb.save(self.file_name)

    @ExcelSheetAutomator.decorator_function
    def format_json(self, excel_file_path, json_file_path):
        """This function will allow us to store the data available in\
        Excel format to json format"""
        read_excel_file = pd.read_excel(excel_file_path, sheet_name='Sheet', index_col=0)
        thisisjson = read_excel_file.to_json(orient='records', indent=2)
        thisisjson_dict = json.loads(thisisjson)

        def file_opener():
            with open(json_file_path, 'w') as json_file:
                json.dump(thisisjson_dict, json_file, indent=2)
        file_opener()
        # read_excel_file.to_json(json_file_path, orient='records', lines=True, indent=2)
        print(f"New Data of {self.file_name} has been updated in the json file. ")

    def open_excel(self):
        """This fucntion ensures that code does not exit unless the cmd is executed."""
        process = subprocess.Popen(["start", "/WAIT", self.file_name], shell=True)
        rc = process.wait()
        if rc == 0:
            print("successfully saved!")
        else:
            pass

    def way_to_trash(self):
        """This function helps us to remove a file temporarily from it's residing\
        location to the recycle bin."""
        try:
            send2trash(self.file_name)
            print(f"{self.file_name} is successfully removed from it's residing location.")
        except FileNotFoundError as msg:
            print(msg)
