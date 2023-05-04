import openpyxl
from openpyxl.styles import Font, PatternFill
from worksheet import SheetLoader
from datetime import datetime


class ExcelDesigner(SheetLoader):

    def __init__(self):
        super().__init__()
        self.load_designs(self.file_name)
        self.empty_filler(self.file_name)
        self.repeat_marker(self.file_name)
        self.open_excel()
        self.format_json(self.file_name, self.path)
        self.way_to_trash()

    def sheet(self, path):
        """This function is to load the worksheet."""
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        return wb, ws

    def load_designs(self, path):
        """This function takes in 1 argument as the path of the file \
            where in you want to load your Excel designs."""
        wb, ws = self.sheet(path)
        ws.row_dimensions[1].height = 80
        for row in range(2, 24):
            ws.row_dimensions[row].height = 30

        ws.column_dimensions['A'].width = 4
        for val in ['B', 'C', 'D']:
            ws.column_dimensions[val].width = 22
        for val in ['F', 'G', 'H', 'I']:
            ws.column_dimensions[val].width = 16

        # defined a style
        font_style = Font(name="Source Sans Pro SemiBold", size=14, color="080600", bold=True)

        # # to apply this font we need to select the cell
        fill_pattern = PatternFill(patternType='solid', fgColor='EDDFAF')
        for col in range(1, 49):
            selected_cell = ws.cell(row=1, column=col)
            selected_cell.font = font_style
            selected_cell.fill = fill_pattern
            wb.save(self.file_name)  # stores it back to the Excel file

    def empty_filler(self, path):
        wb, ws = self.sheet(path)
        rng = ws['F3':'I22']
        fill_pattern = PatternFill(patternType='darkGrid', fgColor='BBBAB8')
        for cell in rng:
            for val in cell:
                if val.value is None:
                    val.fill = fill_pattern
        wb.save(path)

    def repeat_marker(self, path):
        element_lst = []
        wb, ws = self.sheet(path)
        font_style = Font(color="F12D15")
        for val in ws['D']:
            if val.value not in element_lst:
                element_lst.append(val.value)
            else:
                val.font = font_style
        wb.save(path)


def main():
    start_time = datetime.now()
    excel = ExcelDesigner()
    excel.json_reader()
    end_time = datetime.now()
    print(f"Total time taken by the script: {end_time - start_time} seconds")


if __name__ == "__main__":
    main()
